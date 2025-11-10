from influxdb_client import InfluxDBClient
import pandas as pd
from datetime import timedelta
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import pickle
import os

# --- KONFIGURASI INFLUXDB ---
url     = "http://localhost:8086"
token   = "fM-opVTUNZ0FARtHGp5vojc6q2xjddNGgnxnTzMIJ6YOgXAME6i1Wn4LSWVgO1Hx70Vl-UicZu2Am4XsGPgQQA=="
org     = "TEKLIS PNM"
bucket  = "TemperatureSensor"

# --- KONFIGURASI GOOGLE DRIVE ---
SCOPES = ['https://www.googleapis.com/auth/drive.file']
CLIENT_SECRET_FILE = '/home/pi/Documents/Credentials/client_secret_422444898104-5biomledsotgfbsldpimukvkkaveatgl.apps.googleusercontent.com.json'
TOKEN_FILE = '/home/pi/Documents/Credentials/token.pickle'
FOLDER_ID = '1cShAstHWd5sdE6lCjBYCefXTPUNRmOwm'  # ganti dengan folder kamu

# -- AUTENTIKASI GOOGLE DRIVE ---
creds = None
if os.path.exists(TOKEN_FILE):
    with open(TOKEN_FILE, 'rb') as token_file:
        creds = pickle.load(token_file)

if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
        creds = flow.run_local_server(port=0)
    with open(TOKEN_FILE, 'wb') as token_file:
        pickle.dump(creds, token_file)

service = build('drive', 'v3', credentials=creds)

# --- QUERY INFLUXDB ---
query = f'''
from(bucket: "TemperatureSensor")
  |> range(start: -1d)
  |> filter(fn: (r) => r["_measurement"] == "Transformers")
  |> filter(fn: (r) => r["model"] == "TR-01")
  |> pivot(rowKey:["_time"], columnKey: ["_field"], valueColumn: "_value")
  |> keep(columns: ["_time", "surface", "phaseR", "phaseS", "phaseT", "city", "model", "province", "site"])
  |> yield(name: "transformer_testing")
'''

# --- EKPORT DATA DARI INFLUXDB ---
client = InfluxDBClient(url=url, token=token, org=org)
query_api = client.query_api()
tables = query_api.query_data_frame(org=org, query=query)

if tables.empty:
    print("Tidak ada data ditemukan")
    exit()
else:
    # --- Gabungkan semua tabel hasil query ---
    df = tables
    
    # --- HAPUS KOLOM YANG TIDAK DIPERLUKAN ---
    # Hapus kolom 'result' dan 'table' jika ada
    df = df.drop(columns=['result', 'table'], errors='ignore')
    
    # --- ATUR URUTAN KOLOM ---
    df = df[['_time', 'surface', 'phaseR', 'phaseS', 'phaseT', 'model', 'city', 'province', 'site']]
    
    # --- KONVERSI KE WIB ---
    for col in df.columns:
        if "_time" in col or "_start" in col or "_stop" in col:
            try:
                df[col] = pd.to_datetime(df[col], utc=True, errors="coerce")
                df[col] = df[col].dt.tz_convert("Asia/Jakarta")
                df[col] = df[col].dt.tz_localize(None)
                df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S GMT+7")
            except Exception:
                pass  # biarkan kolom lain tetap aman
            
    # --- GANTI NAMA KOLOM ---
    df = df.rename(columns={
        '_time': 'TIME',
        'surface': 'SURFACE',
        'phaseR': 'PHASE R',
        'phaseS': 'PHASE S',
        'phaseT': 'PHASE T',
        'model': 'MODEL',
        'city': 'CITY',
        'province': 'PROVINCE',
        'site': 'SITE'
    })
            
    # --- SIMPAN KE DALAM BENTUK FILE EXCEL ---
    # Pastikan folder penyimpanan ada, jika tidak buat foldernya
    save_dir = "/home/pi/Documents/InfluxDB_Exports/"
    os.makedirs(save_dir, exist_ok=True)
    # Membuat nama file dengan timestamp
    output_file = os.path.join(save_dir, f"influx_export_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    # Simpan dengan RxcelWriter agar bisa di atur formatnya nanti
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ExportedData')
        
        # Ambil workbook dan sheet untuk styling
        workbook = writer.book
        sheet = workbook['ExportedData']
        
        # Buat font bold
        bold_font = Font(bold=True)
        # Terapkan font bold ke header
        for cell in sheet[1]:
            cell.font = bold_font
        
        # AutoFit Lebar Kolom
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # ambil huruf kolom (A, B, C, ...)
            for cell in column_cells:
                try:
                    # Cari panjang teks terpanjang di kolom
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
        
        # Buat tabel otomatis (table style)
        total_rows = sheet.max_row
        total_cols = sheet.max_column
        ref = f"A1:{sheet.cell(row=total_rows, column=total_cols).coordinate}"
        
        table = Table(displayName="ExportedDataTable", ref=ref)
        
        # Gunakan style tabel (zebra stripe)
        style = TableStyleInfo(
            name="TableStyleLight1",
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        sheet.add_table(table)
        
        # freeze Header
        sheet.freeze_panes = "A2"
    
    print(f"Data berhasil diekspor ke: {output_file}")
    
    # --- UPLOAD KE GOOGLE DRIVE ---
    file_metadata = {
        'name': os.path.basename(output_file),
        'parents': [FOLDER_ID]
    }
    media = MediaFileUpload(
        output_file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    uploaded_file = service.files().create(
        body=file_metadata, media_body=media, fields='id'
    ).execute()
    print("File berhasil diunggah ke Google Drive dengan ID:", uploaded_file.get('id'))